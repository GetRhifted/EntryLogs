from django.shortcuts import render, redirect
from django.urls import reverse_lazy, reverse
from django.http import HttpResponse
from django.views.generic import ListView, TemplateView, DetailView, UpdateView, FormView, DeleteView, CreateView
from django.contrib.auth.views import LoginView, LogoutView
from django.db.models import Q
from django.conf import settings
from django.core.exceptions import ValidationError
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.core.exceptions import PermissionDenied
from django.views import View

from django.db.models import Q

from. forms import RegistroGeneralForm, RegistroCanastaForm, RegistroTiemposForm, RegistroDesechosForm, RegistroBrixForm, RegistroEmpacadosForm, CompararRegistrosForm, RegistrodeUsuarioForm, RegistroFresaGeneralForm, RegistroFresaCanastaForm, RegistroFresaTiemposForm, RegistroFresaBrixForm, RegistroFresaEmpacadosForm, CompararRegistrosFresaForm, EditarUsuarioForm
from. models import Registro, RegistroFresa

from formtools.wizard.views import SessionWizardView
import openpyxl
from openpyxl.styles import Border, Side, PatternFill

# Vista de Registro de Canasta de Mora.
class RegistroWizardView(SessionWizardView):
    template_name = 'registros/registro_mora.html'
    form_list = [RegistroGeneralForm, RegistroCanastaForm, RegistroTiemposForm, RegistroDesechosForm,
                 RegistroBrixForm, RegistroEmpacadosForm]

    def done(self, form_list, **kwargs):
        Usuario = self.request.user
        cleaned_data = self.get_all_cleaned_data()
        registro = Registro(Usuario=Usuario, **cleaned_data)
        registro.save()
        return redirect(self.get_success_url())

    def get_success_url(self):
        return reverse('registros:registro_completado_mora')

# Vista que Confirma la creacion del Registro de Canasta de Mora ingresado.    
class RegistroCompletadoView(TemplateView):
    model = Registro
    form_class = RegistroGeneralForm
    template_name = 'registros/registro_completado_mora.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['registro'] = Registro.objects.last()
        return context

    def form_valid(self, form):
        # Guarda el registro y asigna la instancia del objeto a la variable 'registro'
        registro = form.save()
        # Agrega el objeto de registro al contexto de la plantilla
        return render(self.request, self.template_name, {'registro': registro})

# Vista de Creacion de Usuarios.    
class RegistroUsuarioView(SuccessMessageMixin, CreateView):
    template_name = 'registros/registro_usuario.html'
    form_class = RegistrodeUsuarioForm
    success_url = reverse_lazy('registros:home')
    success_message = "Usuario %(username)s creado"

    def form_valid(self, form):
        response = super().form_valid(form)
        username = form.cleaned_data['username']
        messages.success(self.request, self.success_message % {'username': username})
        return response
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        for field in self.form_class.base_fields.values():
            field.help_text = ''


class UserEditView(LoginRequiredMixin, UpdateView):
    model = User
    form_class = EditarUsuarioForm
    template_name = 'registros/editar_usuario.html'
    success_url = reverse_lazy('registros:home')

    def get_object(self, queryset=None):
        return self.request.user
    
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        for field in self.form_class.base_fields.values():
            field.help_text = ''


# Vista de Inicio de Sesion.
class LoginView(LoginView):
    template_name = 'registros/login.html'
    success_url = reverse_lazy('registros:home')

# Vista de Cierre de Sesion.
class MiLogoutView(LogoutView):
    next_page = reverse_lazy('registros:login')

# Vista de Pagina Principal.
class HomeView(ListView):
    template_name = 'registros/home.html'
    context_object_name = 'registros_mora'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['registros_fresa'] = RegistroFresa.objects.all().order_by('-id')[:10]
        return context

    def get_queryset(self):
        return Registro.objects.all().order_by('-id')[:10]
    
# Vista para visualizar los resultados de Busquedas
class ResultadosBusquedaView(ListView):
    template_name = 'registros/resultados_busqueda.html'
    context_object_name = 'registros'

    def get_queryset(self):
        search_query = self.request.GET.get('search_query')
        queryset_mora = list(Registro.objects.filter(Canasta__icontains=search_query))
        queryset_fresa = list(RegistroFresa.objects.filter(Canasta__icontains=search_query))
        queryset = queryset_mora + queryset_fresa
        return queryset
    
# Vista para borrar Registros de Canastas de Mora.
class RegistroDeleteView(DeleteView):
    model = Registro
    template_name = 'registros/borrar_registro_mora.html'
    success_url = reverse_lazy('registros:home')

# Vista para editar las Canastas de Mora.
class RegistroUpdateView(LoginRequiredMixin, UpdateView):
     model = Registro
     form_class = RegistroGeneralForm
     template_name = 'registros/editar_registro_mora.html'

     def dispatch(self, request, *args, **kwargs):
            obj = self.get_object()
            if obj.Usuario != request.user:
                raise PermissionDenied
            return super().dispatch(request, *args, **kwargs)
     
     def get_success_url(self):
        return reverse_lazy('registros:editar_registro_mora2', kwargs={'pk': self.object.pk})
     
class RegistroUpdateView2(LoginRequiredMixin, UpdateView):
     model = Registro
     form_class = RegistroCanastaForm
     template_name = 'registros/editar_registro_mora2.html'

     def dispatch(self, request, *args, **kwargs):
            obj = self.get_object()
            if obj.Usuario != request.user:
                raise PermissionDenied
            return super().dispatch(request, *args, **kwargs)
     
     def get_success_url(self):
        return reverse_lazy('registros:editar_registro_mora3', kwargs={'pk': self.object.pk})
     
class RegistroUpdateView3(LoginRequiredMixin, UpdateView):
     model = Registro
     form_class = RegistroTiemposForm
     template_name = 'registros/editar_registro_mora3.html'

     def dispatch(self, request, *args, **kwargs):
            obj = self.get_object()
            if obj.Usuario != request.user:
                raise PermissionDenied
            return super().dispatch(request, *args, **kwargs)
     
     def get_success_url(self):
        return reverse_lazy('registros:editar_registro_mora4', kwargs={'pk': self.object.pk})

class RegistroUpdateView4(LoginRequiredMixin, UpdateView):
     model = Registro
     form_class = RegistroDesechosForm
     template_name = 'registros/editar_registro_mora4.html'

     def dispatch(self, request, *args, **kwargs):
            obj = self.get_object()
            if obj.Usuario != request.user:
                raise PermissionDenied
            return super().dispatch(request, *args, **kwargs)
     
     def get_success_url(self):
        return reverse_lazy('registros:editar_registro_mora5', kwargs={'pk': self.object.pk})

class RegistroUpdateView5(LoginRequiredMixin, UpdateView):
     model = Registro
     form_class = RegistroBrixForm
     template_name = 'registros/editar_registro_mora5.html'

     def dispatch(self, request, *args, **kwargs):
            obj = self.get_object()
            if obj.Usuario != request.user:
                raise PermissionDenied
            return super().dispatch(request, *args, **kwargs)
     
     def get_success_url(self):
        return reverse_lazy('registros:editar_registro_mora6', kwargs={'pk': self.object.pk})

class RegistroUpdateView6(LoginRequiredMixin, UpdateView):
     model = Registro
     form_class = RegistroEmpacadosForm
     template_name = 'registros/editar_registro_mora6.html'

     def dispatch(self, request, *args, **kwargs):
            obj = self.get_object()
            if obj.Usuario != request.user:
                raise PermissionDenied
            return super().dispatch(request, *args, **kwargs)
     
     def get_success_url(self):
        return reverse_lazy('registros:registro_detallado_mora', kwargs={'pk': self.object.pk})

# Vista Detalle del Registro de Canasta de Mora ingresado.     
class RegistroDetalladoView(DetailView):
    model = Registro
    template_name = 'registros/registro_detallado_mora.html'

# Vista que permite generar el Archivo Excel.
class CompararRegistrosView(FormView):
    template_name = 'registros/comparar_registros.html'
    form_class = CompararRegistrosForm

    def form_valid(self, form):
        registros = form.cleaned_data['registros']
        if registros.count() < settings.NUMERO_MINIMO_REGISTROS:
            form.add_error(
                'registros',
                ValidationError(f'Seleccione al menos {settings.NUMERO_MINIMO_REGISTROS} registros.'),
            )
            return self.form_invalid(form)

        # Obtener los datos de los registros seleccionados
        datos = []
        for registro in registros:
            datos.append([
                registro.Canasta,
                registro.Fecha,
                registro.Contenido_Total,
                registro.Azucar,
                registro.Sorbato,
                registro.Producto_no_Conforme,
                registro.Fruta_Seleccionada,
                registro.Inicio_Jornada,
                registro.Seleccion_del_Producto,
                registro.Mora_Entra_a_la_Olla,
                registro.Hervor_Inicio,
                registro.Hervor_Final,
                registro.Enfriamiento_Inicio,
                registro.Enfriamiento_Final,
                registro.Inicio_Despulpado,
                registro.Final_Despulpado,
                registro.Segunda_Coccion,
                registro.Hora_Final_Mora,
                registro.Inicio_de_Empaque,
                registro.Finalizacion_de_la_Canasta,
                registro.Semilla,
                registro.Pulpa,
                registro.Valor_Primer_Brix,
                registro.Hora_Primer_Brix,
                registro.Valor_Brix_Final,
                registro.Hora_Brix_Final,
                registro.Producto_Terminado,
                registro.Media_Libra,
                registro.Libra,
                registro.Bolsa_Seis_kg,
                registro.Otro,
                registro.Observaciones,
            ])

        # Crear un libro de Excel y una hoja de cálculo
        libro = openpyxl.Workbook()
        hoja = libro.active

        # Escribir los encabezados de la tabla
        encabezados = [
            'Canasta',
            'Fecha',
            'Contenido Total Usado',
            'Azúcar',
            'Sorbato',
            'Producto no Conforme',
            'Fruta Seleccionada',
            'Inicio de la Jornada',
            'Selección del Producto',
            'Hora de puesta en la Olla',
            'Hora de Hervor',
            'Hora de Finalización del Hervor',
            'Inicio del Enfriamiento',
            'Final del Enfriamiento',
            'Inicio del Despulpado',
            'Final del Despulpado',
            'Segunda Cocción',
            'Hora Final de la Mora',
            'Inicio del Empaque',
            'Finalización de la Canasta',
            'Semilla',
            'Pulpa',
            'Valor Primer Brix',
            'Hora Primer Brix',
            'Valor Brix Final',
            'Hora Brix Final',
            'Producto Terminado',
            'Media Libra',
            'Libra',
            'Bolsa Cinco kg',
            'Otro',
            'Observaciones',
        ]
        hoja.append(encabezados)

        # Escribir los datos de los registros
        for fila in datos:
            hoja.append(fila)

        # Crear una tabla con estilo en Excel
        tabla = openpyxl.worksheet.table.Table(displayName="TablaRegistros", ref=f"A1:{openpyxl.utils.get_column_letter(len(encabezados))}{len(datos) + 1}")
        hoja.add_table(tabla)

        # Aplicar bordes a las celdas de la tabla
        borde = Border(
                left=Side(border_style="medium", color="000000"),
                right=Side(border_style="medium", color="000000"),
                top=Side(border_style="medium", color="000000"),
                bottom=Side(border_style="medium", color="000000")
        )
        for row in hoja.iter_rows(min_row=2, max_row=len(datos) + 1, min_col=1, max_col=len(encabezados)):
            for cell in row:
                cell.border = borde
        
        color1 = "F6FDFF"  # Azul
        color2 = "BBDEEA"  # Azul Grisaseo

        # Aplicar los colores intercalados a las filas de la tabla
        fill1 = PatternFill(start_color=color1, end_color=color1, fill_type="solid")
        fill2 = PatternFill(start_color=color2, end_color=color2, fill_type="solid")

        for i, row in enumerate(hoja.iter_rows(min_row=2, max_row=len(datos) + 1, min_col=1, max_col=len(encabezados)), start=1):
            if i % 2 == 0:
                for cell in row:
                    cell.fill = fill1
            else:
                for cell in row:
                    cell.fill = fill2

        # Ajustar el ancho de las columnas
        for columna in hoja.columns:
            max_length = 0
            column = columna[0].column_letter  # Obtiene la letra de la columna
            for cell in columna:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.5  # Ajusta el ancho de la columna
            hoja.column_dimensions[column].width = adjusted_width

        # Crear una respuesta HTTP con el archivo de Excel como contenido
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=RegistrosMora.xlsx'

        # Guardar el libro de Excel en el objeto response
        libro.save(response)

        return response
    

# Vista de Registro de Canasta de Fresa.
class RegistroFresaWizardView(SessionWizardView):
    template_name = 'registros/registro_fresa.html'
    form_list = [RegistroFresaGeneralForm, RegistroFresaCanastaForm, RegistroFresaTiemposForm, RegistroFresaBrixForm,
                 RegistroFresaEmpacadosForm]

    def done(self, form_list, **kwargs):
        Usuario = self.request.user
        cleaned_data = self.get_all_cleaned_data()
        registrofresa = RegistroFresa(Usuario=Usuario, **cleaned_data)
        registrofresa.save()
        return redirect(self.get_success_url())
    
    def get_success_url(self):
        return reverse('registros:registro_completado_fresa')
    
# Vista que Confirma la creacion del Registro de Canasta de Fresa ingresado.    
class RegistroFresaCompletadoView(TemplateView):
    model = RegistroFresa
    form_class = RegistroFresaGeneralForm
    template_name = 'registros/registro_completado_fresa.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['registrofresa'] = RegistroFresa.objects.last()
        return context

    def form_valid(self, form):
        # Guarda el registro y asigna la instancia del objeto a la variable 'registro'
        registrofresa = form.save()
        # Agrega el objeto de registro al contexto de la plantilla
        return render(self.request, self.template_name, {'registrofresa': registrofresa})
    
# Vista Detalle del Registro de Canasta de Mora ingresado.     
class RegistroFresaDetalladoView(DetailView):
    model = RegistroFresa
    template_name = 'registros/registro_detallado_fresa.html'
     
# Vista para editar las Canastas de Fresa.
class RegistroFresaUpdateView(LoginRequiredMixin, UpdateView):
     model = RegistroFresa
     form_class = RegistroFresaGeneralForm
     template_name = 'registros/editar_registro_fresa.html'

     def dispatch(self, request, *args, **kwargs):
            obj = self.get_object()
            if obj.Usuario != request.user:
                raise PermissionDenied
            return super().dispatch(request, *args, **kwargs)
     
     def get_success_url(self):
        return reverse_lazy('registros:editar_registro_fresa2', kwargs={'pk': self.object.pk})
     
class RegistroFresaUpdateView2(LoginRequiredMixin, UpdateView):
     model = RegistroFresa
     form_class = RegistroFresaCanastaForm
     template_name = 'registros/editar_registro_fresa2.html'

     def dispatch(self, request, *args, **kwargs):
            obj = self.get_object()
            if obj.Usuario != request.user:
                raise PermissionDenied
            return super().dispatch(request, *args, **kwargs)
     
     def get_success_url(self):
        return reverse_lazy('registros:editar_registro_fresa3', kwargs={'pk': self.object.pk})
     
class RegistroFresaUpdateView3(LoginRequiredMixin, UpdateView):
     model = RegistroFresa
     form_class = RegistroFresaTiemposForm
     template_name = 'registros/editar_registro_fresa3.html'

     def dispatch(self, request, *args, **kwargs):
            obj = self.get_object()
            if obj.Usuario != request.user:
                raise PermissionDenied
            return super().dispatch(request, *args, **kwargs)
     
     def get_success_url(self):
        return reverse_lazy('registros:editar_registro_fresa4', kwargs={'pk': self.object.pk})

class RegistroFresaUpdateView4(LoginRequiredMixin, UpdateView):
     model = RegistroFresa
     form_class = RegistroFresaBrixForm
     template_name = 'registros/editar_registro_fresa4.html'

     def dispatch(self, request, *args, **kwargs):
            obj = self.get_object()
            if obj.Usuario != request.user:
                raise PermissionDenied
            return super().dispatch(request, *args, **kwargs)
     
     def get_success_url(self):
        return reverse_lazy('registros:editar_registro_fresa5', kwargs={'pk': self.object.pk})

class RegistroFresaUpdateView5(LoginRequiredMixin, UpdateView):
     model = RegistroFresa
     form_class = RegistroFresaEmpacadosForm
     template_name = 'registros/editar_registro_fresa5.html'

     def dispatch(self, request, *args, **kwargs):
            obj = self.get_object()
            if obj.Usuario != request.user:
                raise PermissionDenied
            return super().dispatch(request, *args, **kwargs)
     
     def get_success_url(self):
        return reverse_lazy('registros:registro_detallado_fresa', kwargs={'pk': self.object.pk})

# Vista para borrar Registros de Canastas de Fresa.
class RegistroFresaDeleteView(DeleteView):
    model = RegistroFresa
    template_name = 'registros/borrar_registro_fresa.html'
    success_url = reverse_lazy('registros:home')


# Vista que permite generar el Archivo Excel.
class CompararRegistrosFresaView(FormView):
    template_name = 'registros/comparar_registros_fresa.html'
    form_class = CompararRegistrosFresaForm

    def form_valid(self, form):
        registrosfresa = form.cleaned_data['registrosfresa']
        if registrosfresa.count() < settings.NUMERO_MINIMO_REGISTROS:
            form.add_error(
                'registrosfresa',
                ValidationError(f'Seleccione al menos {settings.NUMERO_MINIMO_REGISTROS} registros.'),
            )
            return self.form_invalid(form)

        # Obtener los datos de los registros seleccionados
        datos = []
        for registro in registrosfresa:
            datos.append([
                registro.Canasta,
                registro.Fecha,
                registro.Contenido_Total,
                registro.Azucar,
                registro.Limon,
                registro.Producto_no_Conforme,
                registro.Fruta_Seleccionada,
                registro.Inicio_Jornada,
                registro.Seleccion_del_Producto,
                registro.Fresa_Entra_a_la_Olla,
                registro.Hervor_Inicio,
                registro.Introduccion_Azucar,
                registro.Introduccion_Limon,
                registro.Hervor_Final,
                registro.Hora_Licuado,
                registro.Hora_Final_Fresa,
                registro.Inicio_de_Empaque,
                registro.Finalizacion_de_la_Canasta,
                registro.Valor_Primer_Brix,
                registro.Hora_Primer_Brix,
                registro.Valor_Brix_Final,
                registro.Hora_Brix_Final,
                registro.Producto_Terminado,
                registro.Media_Libra,
                registro.Libra,
                registro.Bolsa_Seis_kg,
                registro.Otro,
                registro.Observaciones,
            ])

        # Crear un libro de Excel y una hoja de cálculo
        libro = openpyxl.Workbook()
        hoja = libro.active

        # Escribir los encabezados de la tabla
        encabezados = [
            'Canasta',
            'Fecha',
            'Contenido Total Usado',
            'Azúcar',
            'Limón',
            'Producto no Conforme',
            'Fruta Seleccionada',
            'Inicio de la Jornada',
            'Selección del Producto',
            'Hora de puesta en la Olla',
            'Hora de Hervor',
            'Introducción del Azúcar',
            'Introdución del Limón',
            'Hora de Finalización del Hervor',
            'Hora del Licuado',
            'Hora Final de la Fresa',
            'Inicio del Empaque',
            'Finalización de la Canasta',
            'Valor Primer Brix',
            'Hora Primer Brix',
            'Valor Brix Final',
            'Hora Brix Final',
            'Producto Terminado',
            'Media Libra',
            'Libra',
            'Bolsa Cinco kg',
            'Otro',
            'Observaciones',
        ]
        hoja.append(encabezados)

        # Escribir los datos de los registros
        for fila in datos:
            hoja.append(fila)

        # Crear una tabla con estilo en Excel
        tabla = openpyxl.worksheet.table.Table(displayName="TablaRegistros", ref=f"A1:{openpyxl.utils.get_column_letter(len(encabezados))}{len(datos) + 1}")
        hoja.add_table(tabla)

        # Aplicar bordes a las celdas de la tabla
        borde = Border(
                left=Side(border_style="medium", color="000000"),
                right=Side(border_style="medium", color="000000"),
                top=Side(border_style="medium", color="000000"),
                bottom=Side(border_style="medium", color="000000")
        )
        for row in hoja.iter_rows(min_row=2, max_row=len(datos) + 1, min_col=1, max_col=len(encabezados)):
            for cell in row:
                cell.border = borde
        
        color1 = "FFFDFE"  # Azul
        color2 = "FCACC8"  # Azul Grisaseo

        # Aplicar los colores intercalados a las filas de la tabla
        fill1 = PatternFill(start_color=color1, end_color=color1, fill_type="solid")
        fill2 = PatternFill(start_color=color2, end_color=color2, fill_type="solid")

        for i, row in enumerate(hoja.iter_rows(min_row=2, max_row=len(datos) + 1, min_col=1, max_col=len(encabezados)), start=1):
            if i % 2 == 0:
                for cell in row:
                    cell.fill = fill1
            else:
                for cell in row:
                    cell.fill = fill2

        # Ajustar el ancho de las columnas
        for columna in hoja.columns:
            max_length = 0
            column = columna[0].column_letter  # Obtiene la letra de la columna
            for cell in columna:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.5  # Ajusta el ancho de la columna
            hoja.column_dimensions[column].width = adjusted_width

        # Crear una respuesta HTTP con el archivo de Excel como contenido
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=RegistrosFresa.xlsx'

        # Guardar el libro de Excel en el objeto response
        libro.save(response)

        return response
